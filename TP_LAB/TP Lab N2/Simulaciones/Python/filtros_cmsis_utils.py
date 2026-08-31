#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Utilidades compartidas para diseñar filtros y exportarlos a CMSIS-DSP.

Los scripts del TP trabajan internamente en float64, pero exportan exactamente
los coeficientes float32 que utilizarán ``arm_fir_f32`` y
``arm_biquad_cascade_df1_f32``.
"""

from __future__ import annotations

import textwrap
import warnings
from pathlib import Path
from typing import Callable, Mapping, Sequence

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import ScalarFormatter
from scipy import signal


TINY = np.finfo(np.float64).tiny


def magnitud_db(hh: np.ndarray, piso_db: float = -300.0) -> np.ndarray:
    """Convierte una respuesta compleja a dB sin producir log10(0)."""
    piso_lineal = 10.0 ** (piso_db / 20.0)
    return 20.0 * np.log10(np.maximum(np.abs(hh), piso_lineal))


def fase_y_retardo(
    ff_hz: np.ndarray,
    hh: np.ndarray,
    fs_hz: float,
    ocultar_debajo_db: float = -50.0,
) -> tuple[np.ndarray, np.ndarray]:
    """Calcula fase en grados y retardo de grupo numérico en muestras."""
    fase_rad = np.unwrap(np.angle(hh))
    omega = 2.0 * np.pi * np.asarray(ff_hz, dtype=np.float64) / fs_hz
    retardo = -np.gradient(fase_rad, omega)
    mascara = magnitud_db(hh) < ocultar_debajo_db
    # Cerca de un cero de transmisión la fase deja de ser observable y el
    # gradiente numérico genera picos artificiales. También se ocultan dos
    # muestras vecinas de cada lado para no dibujar esos artefactos.
    for _ in range(2):
        mascara = (
            mascara
            | np.r_[False, mascara[:-1]]
            | np.r_[mascara[1:], False]
        )
    retardo[mascara] = np.nan
    return np.rad2deg(fase_rad), retardo


def respuesta_ba(
    bb: np.ndarray,
    aa: np.ndarray | float,
    fs_hz: float,
    puntos: int,
) -> tuple[np.ndarray, np.ndarray]:
    """Respuesta en frecuencia de un filtro dado por numerador/denominador."""
    if puntos < 32:
        raise ValueError("La cantidad de puntos debe ser al menos 32.")
    ff = np.linspace(0.0, fs_hz / 2.0, puntos, endpoint=True)
    return signal.freqz(bb, aa, worN=ff, fs=fs_hz)


def respuesta_sos(
    sos: np.ndarray,
    fs_hz: float,
    puntos: int,
) -> tuple[np.ndarray, np.ndarray]:
    """Respuesta en frecuencia de una cascada de secciones de segundo orden."""
    if puntos < 32:
        raise ValueError("La cantidad de puntos debe ser al menos 32.")
    ff = np.linspace(0.0, fs_hz / 2.0, puntos, endpoint=True)
    if hasattr(signal, "freqz_sos"):
        return signal.freqz_sos(sos, worN=ff, fs=fs_hz)
    return signal.sosfreqz(sos, worN=ff, fs=fs_hz)


def normalizar_sos(sos: np.ndarray) -> np.ndarray:
    """Normaliza cada sección de SciPy para que a0 sea exactamente uno."""
    salida = np.asarray(sos, dtype=np.float64).copy()
    if salida.ndim != 2 or salida.shape[1] != 6:
        raise ValueError("El arreglo SOS debe tener forma (num_etapas, 6).")
    if np.any(np.isclose(salida[:, 3], 0.0)):
        raise ValueError("Hay una sección SOS con a0=0.")

    salida[:, :3] /= salida[:, 3, None]
    salida[:, 4:] /= salida[:, 3, None]
    salida[:, 3] = 1.0
    return salida


def sos_scipy_a_cmsis_df1(sos: np.ndarray) -> np.ndarray:
    """Convierte SOS de SciPy a [b0,b1,b2,-a1,-a2] float32 de CMSIS."""
    sos_n = normalizar_sos(sos)
    cmsis = sos_n[:, [0, 1, 2, 4, 5]].copy()
    cmsis[:, 3:] *= -1.0
    return cmsis.astype(np.float32)


def cmsis_df1_a_sos(cmsis: np.ndarray) -> np.ndarray:
    """Reconstruye SOS para verificar exactamente los coeficientes float32."""
    coeffs = np.asarray(cmsis, dtype=np.float32)
    if coeffs.ndim != 2 or coeffs.shape[1] != 5:
        raise ValueError("Los coeficientes CMSIS deben tener forma (num_etapas, 5).")

    sos = np.empty((coeffs.shape[0], 6), dtype=np.float64)
    sos[:, :3] = coeffs[:, :3]
    sos[:, 3] = 1.0
    sos[:, 4:] = -coeffs[:, 3:]
    return sos


def simular_cmsis_df1_f32(cmsis: np.ndarray, xx: np.ndarray) -> np.ndarray:
    """Referencia directa de la ecuación usada por el biquad DF1 de CMSIS."""
    coeffs = np.asarray(cmsis, dtype=np.float32)
    estado = np.zeros((coeffs.shape[0], 4), dtype=np.float32)
    yy = np.empty(np.asarray(xx).size, dtype=np.float32)

    for nn, muestra in enumerate(np.asarray(xx, dtype=np.float32)):
        entrada = np.float32(muestra)
        for etapa, (b0, b1, b2, a1, a2) in enumerate(coeffs):
            x1, x2, y1, y2 = estado[etapa]
            salida = np.float32(
                np.float32(b0 * entrada)
                + np.float32(b1 * x1)
                + np.float32(b2 * x2)
                + np.float32(a1 * y1)
                + np.float32(a2 * y2)
            )
            estado[etapa] = (entrada, x1, salida, y1)
            entrada = salida
        yy[nn] = entrada
    return yy


def verificar_conversion_cmsis(sos_arm: np.ndarray, cmsis: np.ndarray) -> float:
    """Compara la respuesta al impulso SciPy contra la ecuación CMSIS f32."""
    impulso = np.zeros(8192, dtype=np.float32)
    impulso[0] = 1.0
    y_scipy = signal.sosfilt(np.asarray(sos_arm, dtype=np.float32), impulso)
    y_cmsis = simular_cmsis_df1_f32(cmsis, impulso)
    error = float(np.max(np.abs(y_scipy - y_cmsis)))
    if error > 2.0e-5:
        raise RuntimeError(
            f"La conversión SciPy/CMSIS no superó la verificación: {error:.3e}"
        )
    return error


def verificar_estabilidad_sos(sos: np.ndarray) -> float:
    """Devuelve el radio máximo de los polos y rechaza filtros inestables."""
    _, polos, _ = signal.sos2zpk(sos)
    radio_max = float(np.max(np.abs(polos))) if polos.size else 0.0
    if radio_max >= 1.0:
        raise RuntimeError(f"Filtro inestable después de cuantizar: |p|max={radio_max}")
    return radio_max


def _float32_c(valor: float, digitos_significativos: int) -> str:
    """Nueve cifras significativas alcanzan para round-trip de float32."""
    if digitos_significativos < 9:
        raise ValueError("Use al menos 9 cifras significativas para float32.")
    decimales = digitos_significativos - 1
    return f"{float(np.float32(valor)):+.{decimales}e}f"


NOMBRE_EXCEL_COEFICIENTES = "coeficientes_filtros_STM32.xlsx"
ORDEN_HOJAS_EXCEL = (
    "FIR Notch",
    "IIR Notch",
    "IIR Chebyshev",
    "FIR Pasabajos",
)


def _nombre_hoja(nombre: str) -> str:
    nombres = {
        "fir_notch": "FIR Notch",
        "iir_notch": "IIR Notch",
        "iir_cheby": "IIR Chebyshev",
        "fir_pasabajos": "FIR Pasabajos",
    }
    return nombres.get(nombre, nombre.replace("_", " ").title())[:31]


def _abrir_excel(ruta_excel: Path):
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Falta openpyxl. Ejecute: python -m pip install -r requirements.txt"
        ) from exc

    if ruta_excel.exists():
        return load_workbook(ruta_excel)
    libro = Workbook()
    libro.active.title = "_temporal"
    return libro


def _guardar_hoja_coeficientes(
    ruta_excel: Path,
    nombre: str,
    descripcion: str,
    fs_hz: float,
    filas_python: Sequence[tuple[int, int | str, str, float]],
    filas_stm32: Sequence[tuple[int, int | str, str, float]],
    declaracion_c: str,
    nota_orden: str,
    digitos_c: int,
) -> Path:
    """Actualiza una hoja del único Excel consolidado de coeficientes."""
    ruta_excel.parent.mkdir(parents=True, exist_ok=True)
    libro = _abrir_excel(ruta_excel)
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    hoja_nombre = _nombre_hoja(nombre)
    if hoja_nombre in libro.sheetnames:
        libro.remove(libro[hoja_nombre])
    hoja = libro.create_sheet(hoja_nombre)

    azul = "1F4E78"
    celeste = "D9EAF7"
    verde = "E2F0D9"
    gris = "E7E6E6"
    blanco = "FFFFFF"
    borde_inferior = Border(bottom=Side(style="thin", color="A6A6A6"))

    hoja.sheet_view.showGridLines = False
    hoja.freeze_panes = "A7"
    hoja.merge_cells("A1:E1")
    hoja["A1"] = f"Coeficientes de {descripcion}"
    hoja["A1"].font = Font(name="Aptos Display", size=15, bold=True, color=blanco)
    hoja["A1"].fill = PatternFill("solid", fgColor=azul)
    hoja["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    hoja.row_dimensions[1].height = 38
    hoja["A2"] = "Frecuencia de muestreo"
    hoja["B2"] = fs_hz
    hoja["C2"] = "Hz"
    hoja["A2"].font = Font(bold=True)
    hoja["B2"].number_format = "0.###############"

    fila = 4
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    hoja.cell(fila, 1, "Coeficientes calculados por Python (float64)")
    hoja.cell(fila, 1).fill = PatternFill("solid", fgColor=celeste)
    hoja.cell(fila, 1).font = Font(bold=True, color="17365D")

    fila += 1
    encabezados = ["Índice", "Etapa", "Nombre", "Valor numérico", "Decimal exacto"]
    for columna, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(fila, columna, texto)
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor=gris)
        celda.border = borde_inferior

    for indice, etapa, etiqueta, valor in filas_python:
        fila += 1
        hoja.cell(fila, 1, indice)
        hoja.cell(fila, 2, etapa)
        hoja.cell(fila, 3, etiqueta)
        hoja.cell(fila, 4, float(valor))
        hoja.cell(fila, 4).number_format = "0.00000000000000000E+00"
        hoja.cell(fila, 5, f"{float(valor):+.17e}")
        hoja.cell(fila, 5).number_format = "@"

    fila += 2  # deja una fila completamente libre entre ambas secciones
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    hoja.cell(fila, 1, "Coeficientes para STM32 / CMSIS-DSP")
    hoja.cell(fila, 1).fill = PatternFill("solid", fgColor=verde)
    hoja.cell(fila, 1).font = Font(bold=True, color="375623")

    fila += 1
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    hoja.cell(fila, 1, nota_orden)
    hoja.cell(fila, 1).alignment = Alignment(wrap_text=True)

    fila += 1
    for columna, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(fila, columna, texto)
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor=gris)
        celda.border = borde_inferior

    for indice, etapa, etiqueta, valor in filas_stm32:
        fila += 1
        hoja.cell(fila, 1, indice)
        hoja.cell(fila, 2, etapa)
        hoja.cell(fila, 3, etiqueta)
        hoja.cell(fila, 4, float(np.float32(valor)))
        hoja.cell(fila, 4).number_format = "0.00000000000000000E+00"
        hoja.cell(fila, 5, _float32_c(valor, digitos_c))
        hoja.cell(fila, 5).number_format = "@"

    fila += 2
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    hoja.cell(fila, 1, "Bloque listo para copiar y pegar en app.c")
    hoja.cell(fila, 1).fill = PatternFill("solid", fgColor=azul)
    hoja.cell(fila, 1).font = Font(bold=True, color=blanco)

    fila += 1
    inicio_codigo = fila
    hoja.cell(fila, 1, declaracion_c + " = {")
    for _, _, _, valor in filas_stm32:
        fila += 1
        hoja.cell(fila, 1, f"    {_float32_c(valor, digitos_c)},")
    fila += 1
    hoja.cell(fila, 1, "};")
    for numero_fila in range(inicio_codigo, fila + 1):
        hoja.cell(numero_fila, 1).font = Font(name="Consolas", size=10)
        hoja.cell(numero_fila, 1).number_format = "@"

    hoja.column_dimensions["A"].width = 22
    hoja.column_dimensions["B"].width = 14
    hoja.column_dimensions["C"].width = 22
    hoja.column_dimensions["D"].width = 28
    hoja.column_dimensions["E"].width = 31
    hoja.auto_filter.ref = f"A5:E{5 + len(filas_python)}"

    if "_temporal" in libro.sheetnames:
        libro.remove(libro["_temporal"])
    orden = {nombre_hoja: idx for idx, nombre_hoja in enumerate(ORDEN_HOJAS_EXCEL)}
    libro._sheets.sort(key=lambda item: orden.get(item.title, len(orden)))

    try:
        libro.save(ruta_excel)
    except PermissionError as exc:
        raise RuntimeError(
            f"Cierre el Excel antes de ejecutar el programa: {ruta_excel}"
        ) from exc
    return ruta_excel


def exportar_fir_cmsis(
    carpeta: Path,
    nombre: str,
    bb: np.ndarray,
    fs_hz: float,
    descripcion: str,
    digitos_c: int = 9,
) -> Mapping[str, Path]:
    """Agrega un FIR al único Excel con datos Python y datos para CMSIS."""
    bb64 = np.asarray(bb, dtype=np.float64)
    bb_cmsis = bb64[::-1].astype(np.float32)
    filas_python = [
        (indice, "FIR", f"b[{indice}]", float(valor))
        for indice, valor in enumerate(bb64)
    ]
    filas_stm32 = [
        (indice, "FIR", f"tap_invertido[{indice}]", float(valor))
        for indice, valor in enumerate(bb_cmsis)
    ]
    ruta_excel = carpeta.parent / NOMBRE_EXCEL_COEFICIENTES
    _guardar_hoja_coeficientes(
        ruta_excel=ruta_excel,
        nombre=nombre,
        descripcion=descripcion,
        fs_hz=fs_hz,
        filas_python=filas_python,
        filas_stm32=filas_stm32,
        declaracion_c="float32_t float_fir_taps[FIR_TAP_NUM]",
        nota_orden=(
            "Los taps están cuantizados a float32 y en orden temporal inverso, "
            "tal como los espera arm_fir_f32."
        ),
        digitos_c=digitos_c,
    )
    return {"excel": ruta_excel}


def exportar_sos_cmsis(
    carpeta: Path,
    nombre: str,
    sos: np.ndarray,
    fs_hz: float,
    descripcion: str,
    digitos_c: int = 9,
) -> tuple[Mapping[str, Path], np.ndarray, np.ndarray]:
    """Agrega un IIR al Excel y devuelve su representación CMSIS float32."""
    sos64 = normalizar_sos(sos)
    cmsis = sos_scipy_a_cmsis_df1(sos64)
    sos_arm = cmsis_df1_a_sos(cmsis)
    nombres_python = ("b0", "b1", "b2", "a0", "a1", "a2")
    nombres_cmsis = ("b0", "b1", "b2", "-a1", "-a2")
    filas_python = []
    indice = 0
    for etapa, fila in enumerate(sos64, start=1):
        for etiqueta, valor in zip(nombres_python, fila):
            filas_python.append((indice, etapa, etiqueta, float(valor)))
            indice += 1
    filas_stm32 = []
    indice = 0
    for etapa, fila in enumerate(cmsis, start=1):
        for etiqueta, valor in zip(nombres_cmsis, fila):
            filas_stm32.append((indice, etapa, etiqueta, float(valor)))
            indice += 1

    ruta_excel = carpeta.parent / NOMBRE_EXCEL_COEFICIENTES
    _guardar_hoja_coeficientes(
        ruta_excel=ruta_excel,
        nombre=nombre,
        descripcion=descripcion,
        fs_hz=fs_hz,
        filas_python=filas_python,
        filas_stm32=filas_stm32,
        declaracion_c="float32_t float_iir_taps[IIR_TAP_NUM]",
        nota_orden=(
            "Orden por etapa: b0, b1, b2, -a1, -a2. Los signos de "
            "realimentación ya están convertidos para arm_biquad_cascade_df1_f32."
        ),
        digitos_c=digitos_c,
    )
    return {"excel": ruta_excel}, cmsis, sos_arm


def imprimir_coeficientes_fir(
    bb: np.ndarray,
    digitos_c: int = 17,
    coeficientes_por_linea: int = 8,
) -> None:
    """Muestra solamente el bloque FIR de CMSIS listo para pegar en app.c."""
    if coeficientes_por_linea < 1:
        raise ValueError("coeficientes_por_linea debe ser mayor o igual que 1.")

    bb64 = np.asarray(bb, dtype=np.float64)
    bb_cmsis = bb64[::-1].astype(np.float32)
    print("\nCOEFICIENTES PARA STM32/CMSIS — COPIAR Y PEGAR")
    print("float32_t float_fir_taps[FIR_TAP_NUM] = {")
    for inicio in range(0, bb_cmsis.size, coeficientes_por_linea):
        grupo = bb_cmsis[inicio : inicio + coeficientes_por_linea]
        texto = ", ".join(_float32_c(valor, digitos_c) for valor in grupo)
        print(f"    {texto},")
    print("};")


def imprimir_coeficientes_iir(
    sos: np.ndarray,
    cmsis: np.ndarray,
    digitos_c: int = 17,
) -> None:
    """Muestra una vez los SOS de Python y el bloque CMSIS para app.c."""
    sos64 = normalizar_sos(sos)
    print("\nCOEFICIENTES CALCULADOS POR PYTHON")
    print("Orden por etapa: b0, b1, b2, a0, a1, a2")
    print(np.array2string(sos64, precision=17, suppress_small=False))
    print("\nCOEFICIENTES PARA STM32/CMSIS — COPIAR Y PEGAR")
    print("Orden por etapa: b0, b1, b2, -a1, -a2")
    print("float32_t float_iir_taps[IIR_TAP_NUM] = {")
    for valor in np.asarray(cmsis, dtype=np.float32).ravel():
        print(f"    {_float32_c(valor, digitos_c)},")
    print("};")


def graficar_polos_ceros(
    ceros: np.ndarray,
    polos: np.ndarray,
    titulo: str,
    ruta_png: Path | None = None,
    mostrar: bool = True,
    multiplicidad_polo_origen: int = 0,
    nota: str | None = None,
) -> plt.Figure:
    """Dibuja el plano Z, el círculo unitario y los polos/ceros indicados."""
    ceros = np.asarray(ceros, dtype=np.complex128)
    polos = np.asarray(polos, dtype=np.complex128)
    fig, ax = plt.subplots(figsize=(7.5, 7.5))
    angulo = np.linspace(0.0, 2.0 * np.pi, 720)
    ax.plot(np.cos(angulo), np.sin(angulo), "k--", lw=1.2, label="círculo unitario")
    ax.axhline(0.0, color="0.55", lw=0.8)
    ax.axvline(0.0, color="0.55", lw=0.8)

    if ceros.size:
        ax.scatter(
            ceros.real,
            ceros.imag,
            s=70,
            facecolors="none",
            edgecolors="tab:blue",
            linewidths=1.6,
            label="ceros",
        )
    if polos.size:
        ax.scatter(
            polos.real,
            polos.imag,
            s=70,
            marker="x",
            color="tab:red",
            linewidths=1.8,
            label="polos",
        )
    if multiplicidad_polo_origen > 0:
        ax.scatter(
            [0.0],
            [0.0],
            s=80,
            marker="x",
            color="tab:red",
            linewidths=1.8,
            label="polos en el origen",
        )
        ax.annotate(
            f"×{multiplicidad_polo_origen}",
            xy=(0.0, 0.0),
            xytext=(0.06, -0.10),
            textcoords="data",
            fontsize=9,
            color="tab:red",
        )

    radios = np.abs(np.r_[ceros, polos]) if ceros.size or polos.size else np.asarray([1.0])
    limite = max(1.15, 1.12 * float(np.max(radios)))
    ax.set_xlim(-limite, limite)
    ax.set_ylim(-limite, limite)
    ax.set_aspect("equal", adjustable="box")
    ax.set_xlabel("Parte real")
    ax.set_ylabel("Parte imaginaria")
    ax.set_title(titulo, fontsize=12, wrap=True)
    ax.grid(True, which="both", alpha=0.4)
    ax.legend(loc="upper left")
    if nota:
        nota_envuelta = "\n".join(textwrap.wrap(nota, width=82))
        fig.text(
            0.5,
            0.025,
            nota_envuelta,
            ha="center",
            fontsize=8.5,
            va="center",
            bbox={"boxstyle": "round", "facecolor": "white", "alpha": 0.8},
        )
    fig.tight_layout(rect=(0.0, 0.10 if nota else 0.0, 1.0, 1.0))
    if ruta_png is not None:
        ruta_png.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(ruta_png, dpi=180, bbox_inches="tight")
    if mostrar:
        plt.show()
    return fig


def graficar_polos_ceros_sos(
    sos: np.ndarray,
    titulo: str,
    ruta_png: Path | None = None,
    mostrar: bool = True,
) -> plt.Figure:
    ceros, polos, _ = signal.sos2zpk(np.asarray(sos, dtype=np.float64))
    return graficar_polos_ceros(ceros, polos, titulo, ruta_png, mostrar)


def graficar_polos_ceros_fir(
    bb: np.ndarray,
    titulo: str,
    ruta_png: Path | None = None,
    mostrar: bool = True,
    ceros_relevantes: Sequence[complex] | None = None,
    maximo_raices_completas: int = 256,
) -> plt.Figure:
    """Grafica todas las raíces de un FIR corto o una vista estable del largo."""
    bb = np.asarray(bb, dtype=np.float64)
    orden = bb.size - 1
    nota = None
    if orden <= maximo_raices_completas:
        # Algunos cortes exactos producen taps extremos prácticamente nulos.
        # tf2zpk los elimina antes de hallar raíces; se oculta solamente la
        # advertencia de acondicionamiento porque esa reducción es intencional.
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", signal.BadCoefficients)
            ceros, _, _ = signal.tf2zpk(bb, [1.0])
    elif ceros_relevantes is not None:
        ceros = np.asarray(ceros_relevantes, dtype=np.complex128)
        nota = (
            f"Vista estructural: se muestran los ceros especificados. El FIR tiene "
            f"orden {orden}; calcular y dibujar sus {orden} raíces completas no es "
            "numéricamente recomendable."
        )
    else:
        raise ValueError(
            "Para un FIR largo indique ceros_relevantes para el diagrama estable."
        )
    return graficar_polos_ceros(
        ceros=ceros,
        polos=np.asarray([], dtype=np.complex128),
        titulo=titulo,
        ruta_png=ruta_png,
        mostrar=mostrar,
        multiplicidad_polo_origen=orden,
        nota=nota,
    )


def graficar_respuestas(
    respuestas: Mapping[str, tuple[np.ndarray, np.ndarray]],
    fs_hz: float,
    titulo: str,
    dibujar_plantilla: Callable[[plt.Axes], None],
    puntos_xlim: tuple[float, float] | None = None,
    modulo_ylim: tuple[float, float] | None = None,
    ruta_png: Path | None = None,
    mostrar: bool = True,
) -> plt.Figure:
    """Grafica módulo, fase y retardo con el estilo usado en la cátedra."""
    fig, axes = plt.subplots(3, 1, figsize=(11, 10), sharex=True)
    ax_mag, ax_fase, ax_retardo = axes
    retardos_validos: list[np.ndarray] = []
    fases_trazadas: list[tuple[np.ndarray, np.ndarray]] = []
    retardos_trazados: list[tuple[np.ndarray, np.ndarray]] = []

    for etiqueta, (ff, hh) in respuestas.items():
        fase, retardo = fase_y_retardo(ff, hh, fs_hz)
        valores_validos = retardo[np.isfinite(retardo)]
        if valores_validos.size:
            retardos_validos.append(valores_validos)
        fases_trazadas.append((ff, fase))
        retardos_trazados.append((ff, retardo))
        ax_mag.plot(ff, magnitud_db(hh), label=etiqueta)
        ax_fase.plot(ff, fase, label=etiqueta)
        ax_retardo.plot(ff, retardo, label=etiqueta)

    dibujar_plantilla(ax_mag)
    ax_mag.set_title(titulo)
    ax_mag.set_ylabel("Módulo [dB]")
    ax_fase.set_ylabel("Fase [grados]")
    ax_retardo.set_ylabel("Retardo [muestras]")
    ax_retardo.set_xlabel("Frecuencia [Hz]")
    ax_retardo.yaxis.set_major_formatter(ScalarFormatter(useOffset=False))

    if retardos_validos:
        todos_los_retardos = np.concatenate(retardos_validos)
        percentil_bajo, percentil_alto = np.percentile(
            todos_los_retardos, [1.0, 99.0]
        )
        centro = float(np.median(todos_los_retardos))
        escala = max(1.0, abs(centro))
        if percentil_alto - percentil_bajo < 1e-4 * escala:
            # Evita el poco claro "1e-5 + 2e3" para un FIR cuyo retardo
            # teórico es constante e igual a 2000 muestras.
            ax_retardo.set_ylim(centro - 1.0, centro + 1.0)

    if puntos_xlim is not None:
        ax_retardo.set_xlim(*puntos_xlim)
        f_min, f_max = puntos_xlim

        def valores_en_ventana(
            trazas: list[tuple[np.ndarray, np.ndarray]],
        ) -> np.ndarray:
            partes = []
            for ff, valores in trazas:
                mascara = (
                    (ff >= f_min)
                    & (ff <= f_max)
                    & np.isfinite(valores)
                )
                if np.any(mascara):
                    partes.append(valores[mascara])
            return np.concatenate(partes) if partes else np.asarray([])

        fase_visible = valores_en_ventana(fases_trazadas)
        if fase_visible.size:
            fase_min = float(np.min(fase_visible))
            fase_max = float(np.max(fase_visible))
            margen_fase = max(5.0, 0.04 * (fase_max - fase_min))
            ax_fase.set_ylim(fase_min - margen_fase, fase_max + margen_fase)

        retardo_visible = valores_en_ventana(retardos_trazados)
        if retardo_visible.size:
            ret_min = float(np.min(retardo_visible))
            ret_max = float(np.max(retardo_visible))
            if ret_max - ret_min < 1e-4 * max(1.0, abs(float(np.median(retardo_visible)))):
                centro_visible = float(np.median(retardo_visible))
                ax_retardo.set_ylim(centro_visible - 1.0, centro_visible + 1.0)
            else:
                margen_retardo = max(0.1, 0.04 * (ret_max - ret_min))
                ax_retardo.set_ylim(
                    min(0.0, ret_min - margen_retardo),
                    ret_max + margen_retardo,
                )
    if modulo_ylim is not None:
        ax_mag.set_ylim(*modulo_ylim)

    for ax in axes:
        ax.grid(True, which="both", alpha=0.4)
        ax.legend(loc="best")

    fig.tight_layout()
    if ruta_png is not None:
        ruta_png.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(ruta_png, dpi=180, bbox_inches="tight")
    if mostrar:
        plt.show()
    return fig


def cruce_lineal(
    x1: float, y1: float, x2: float, y2: float, nivel: float
) -> float:
    """Interpolación lineal de la abscisa correspondiente a ``nivel``."""
    if np.isclose(y1, y2):
        return 0.5 * (x1 + x2)
    return float(x1 + (nivel - y1) * (x2 - x1) / (y2 - y1))


def medir_notch(
    ff: np.ndarray,
    hh: np.ndarray,
    f0_hz: float,
    nivel_bw_db: float = -3.0,
) -> Mapping[str, float]:
    """Mide frecuencia mínima, profundidad y cruces alrededor del notch."""
    mag = magnitud_db(hh)
    centro = int(np.argmin(np.abs(ff - f0_hz)))
    ancho_busqueda = max(4, ff.size // 5)
    i0 = max(0, centro - ancho_busqueda)
    i1 = min(ff.size, centro + ancho_busqueda + 1)
    minimo = i0 + int(np.argmin(mag[i0:i1]))

    candidatos_izq = np.where(mag[:minimo] >= nivel_bw_db)[0]
    candidatos_der = np.where(mag[minimo + 1 :] >= nivel_bw_db)[0]
    if candidatos_izq.size == 0 or candidatos_der.size == 0:
        return {
            "f_notch_hz": float(ff[minimo]),
            "profundidad_db": float(mag[minimo]),
            "f_3db_izq_hz": np.nan,
            "f_3db_der_hz": np.nan,
            "bw_3db_hz": np.nan,
        }

    izq = int(candidatos_izq[-1])
    der = int(minimo + 1 + candidatos_der[0])
    f_izq = cruce_lineal(ff[izq], mag[izq], ff[izq + 1], mag[izq + 1], nivel_bw_db)
    f_der = cruce_lineal(ff[der - 1], mag[der - 1], ff[der], mag[der], nivel_bw_db)
    return {
        "f_notch_hz": float(ff[minimo]),
        "profundidad_db": float(mag[minimo]),
        "f_3db_izq_hz": f_izq,
        "f_3db_der_hz": f_der,
        "bw_3db_hz": float(f_der - f_izq),
    }


def imprimir_rutas(rutas: Mapping[str, Path]) -> None:
    for tipo, ruta in rutas.items():
        print(f"{tipo}: {ruta.resolve()}")
